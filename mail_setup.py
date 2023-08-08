import logging
import time, random, json, openpyxl, platform, os, shutil
from unicodedata import name
from datetime import datetime
from selenium import webdriver
from random import randint
from openpyxl import load_workbook
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoAlertPresentException, TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.alert import Alert

chrome_options = Options()
chrome_options.add_argument("--start-maximized")

current_folder_path = os.path.dirname(os.path.realpath(__file__))
system_name = str(platform.system())
if system_name == "Windows":
    slash = "\\"
    user_repository = current_folder_path.split("AppData")[0]
else:
    slash = "/"
    user_repository = current_folder_path.replace("MailAnalysis", "")

json_file = current_folder_path + "%smail_config.json" % slash
with open(json_file) as json_data_file:
    data = json.load(json_data_file)

class Objects():
    current_folder_path = os.path.dirname(os.path.realpath(__file__))
    folder_execution = "C:\\Users\\Hanbiro"

    '''if "TestFolder" in current_folder_path:
        myfolder = True
    else:
        myfolder = False'''

    myfolder = True
    global slash

    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    today = now.strftime("%Y/%m/%d")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]

    chromedriver_file = current_folder_path + "%schromedriver.exe" % slash
    
    log_folder = current_folder_path + slash + "Log" + slash
    assets_folder = current_folder_path + slash + "assets" + slash
    log_all = "%smail-log-%s.txt" % (log_folder, date_id)
    log_msg = "%smail-msg-log.txt" % log_folder
    excel_current_section = "%scurrent_section_data.xlsx" % assets_folder

class Files():
    def ConfigFiles():
        if system_name == "Windows" and Objects.myfolder == False:
            config_files = ["chromedriver.exe",
                            "mail_setup.py", 
                            "mail_functions.py",
                            "mail_ui.py",
                            "mail_config.json"]
            for config_file in config_files:
                source_path = Objects.user_repository + "MailAnalysis" + slash + config_file
                destination_path = Objects.current_folder_path + slash + config_file
                shutil.copy(source_path, destination_path)

            #folder_source_path = Objects.
            #  + slash + "MailAnalysis" + slash + "Attachment"
            #folder_destination_path = Objects.current_folder_path + slash + "Attachment"
            #shutil.copytree(folder_source_path, folder_destination_path)

class Logs():
    def CreateLogFiles():
        all_logs = open(Objects.log_all, "w")
        all_logs.close()

        msg_logs = open(Objects.log_msg, "w")
        msg_logs.close() 
    
    def ClearExcelFile():
        wb = load_workbook(Objects.excel_current_section)
        current_sheet = wb.active

        last_row = current_sheet.max_row
        if last_row > 1:
            current_sheet.delete_rows(2, last_row)

        wb.save(Objects.excel_current_section)

    def Logging(msg):
        print(msg)
        log_msg = open(Objects.log_all, "a")
        written_msg = str(msg).encode(encoding="ascii",errors="ignore")
        logged_msg = str(written_msg, "utf-8") + "\n"
        log_msg.write(logged_msg)
        log_msg.close()
    
    def MsgLogging(msg):
        Logs.Logging(msg)
        log_msg = open(Objects.log_msg, "a")
        written_msg = str(msg).encode(encoding="ascii",errors="ignore")
        logged_msg = str(written_msg, "utf-8") + "\n"
        log_msg.write(logged_msg)
        log_msg.close()
    
    def WriteInExcel(**this_dict):
        wb = load_workbook(Objects.excel_current_section)
        current_sheet = wb.active
        current_row = current_sheet.max_row + 1

        for key in this_dict.keys():
            column = this_dict[key]["column"]
            cell_value = this_dict[key]["cell_value"]
            current_sheet.cell(row=current_row, column=column).value = cell_value

        # Column [1]: Title
        # Column [2]: Folder Name
        # Column [3]: Page
        # Column [4]: Position
        # Column [5]: Important mails
        # Column [6]: Mail Type 
        #       (suspected_mails, important_mails, groupware_mails, other_mails)
        # Column [7]: Groupware Mails
        # Column [8]: Menu
        # Column [9]: Status
        # Column [10]: Send Date
        # Column [11]: Folder ID
        # Column [12]: List Total
        # Column [13]: Checkbox

        wb.save(Objects.excel_current_section)

    def WriteInExcel_Checkbox(selected_text):
        wb = load_workbook(Objects.excel_current_section)
        current_sheet = wb.active
        last_row = current_sheet.max_row

        for row in range(1, last_row):
            row+=1
            
            mail_title_value = current_sheet.cell(row=row, column=1).value
            selected_mail_title = str(selected_text).split(" [Date: ")[0]
            print("mail_title_value: -> " + str(mail_title_value))
            print("selected_mail_title: -> " + str(selected_mail_title))
            if mail_title_value == selected_mail_title:
                current_sheet.cell(row=row, column=13).value = True
                break
        
        wb.save(Objects.excel_current_section)

    def CollectExcelList(list_type):
        print("hello : 1")
        wb = load_workbook(Objects.excel_current_section)
        current_sheet = wb.active
        last_row = current_sheet.max_row
        print(" last_row :",last_row)
        
        mail_dict = {}

        for row in range(1, last_row):
            row+=1
            
            mail_title = current_sheet.cell(row=row, column=1).value
            folder_name = current_sheet.cell(row=row, column=2).value
            page_number = current_sheet.cell(row=row, column=3).value
            page_index = current_sheet.cell(row=row, column=4).value
            isImportant = current_sheet.cell(row=row, column=5).value
            mail_type = current_sheet.cell(row=row, column=6).value
            isGroupware = current_sheet.cell(row=row, column=7).value
            groupware_menu = current_sheet.cell(row=row, column=8).value
            resolved_status = current_sheet.cell(row=row, column=9).value
            send_date = current_sheet.cell(row=row, column=10).value
            folder_id = current_sheet.cell(row=row, column=11).value
            list_total = current_sheet.cell(row=row, column=12).value
            mail_checkbox = current_sheet.cell(row=row, column=13).value

            print(" list_type :",list_type)
            print(" mail_type :",mail_type)
            if mail_type == list_type:
                mail_dict_key = "%s [Date: %s]" % (mail_title, send_date)
                mail_dict.update({mail_dict_key: {}})
                mail_dict[mail_dict_key].update({
                        "title": mail_title,
                        "folder_name": folder_name,
                        "page_number": page_number,
                        "page_index": page_index,
                        "isImportant": isImportant,
                        "mail_type": mail_type,
                        "isGroupware": isGroupware,
                        "groupware_menu": groupware_menu,
                        "resolved_status": resolved_status,
                        "send_date": send_date,
                        "folder_id": folder_id,
                        "list_total": list_total,
                        "mail_checkbox": mail_checkbox
                })
        
        return mail_dict

class Driver():
    def StartWebdriver():
        global driver

        args = ["hide_console"]
        driver = webdriver.Chrome(Objects.chromedriver_file, service_args=args)

        try:
            driver.maximize_window()
        except:
            Files.ConfigFiles()
            driver = webdriver.Chrome(Objects.chromedriver_file, service_args=args)
            driver.maximize_window()
        
        return driver

class Waits():
    def WaitElementLoaded(time, xpath):
        '''• Usage: Wait until element VISIBLE in a selected time period'''
        
        WebDriverWait(driver, time).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)

        return element
    
    def Wait10s_ElementClickable(xpath):
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)

        return element

    def Wait10s_ElementLoaded(xpath):
        '''• Usage: Wait 10s until element VISIBLE'''
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)

        return element

    def WaitElementInvisibility(time, xpath):
        '''• Usage: Wait until element INVISIBLE in a selected time period'''
        
        WebDriverWait(driver, time).until(EC.invisibility_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)

        return element

    def Wait10s_ElementInvisibility(xpath):
        '''• Usage: Wait 10s until element INVISIBLE'''
        
        i=0
        for i in range(0,10):
            i+=1
            time.sleep(1)
            try:
                driver.find_element_by_xpath(xpath)
            except WebDriverException:
                break
    
    def WaitUntilPageIsLoaded(page_xpath):
        if bool(page_xpath) == True:
            # wait until page's element is present
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, page_xpath)))

        # check if the loading icon is not present at the page -> page is completely loaded
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='loading-dialog hide']")))
        except WebDriverException:
            pass

        '''If page_xpath=None/False -> only check if the loading icon is not present'''

class Commands():
    def FindElement(xpath):
        element = driver.find_element_by_xpath(xpath)

        return element

    def FindElements(xpath):
        element = driver.find_elements_by_xpath(xpath)

        return element

    def ClickElement(xpath):
        '''• Usage: Do the click on element
                return WebElement'''

        element = driver.find_element_by_xpath(xpath)
        element.click()

        return element

    def ClickElements(xpath, element_position):
        '''• Usage: Do the click on element
                return WebElement'''

        element = driver.find_elements_by_xpath(xpath)
        time.sleep(1)
        element[element_position].click()

        return element

    def Wait10s_ClickElement(xpath):
        '''• Usage: Wait until the element visible and do the click
                return WebElement'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.click()

        return element

    def InputElement(xpath, value):
        '''• Usage: Send key value in input box
                return WebElement'''

        element = driver.find_element_by_xpath(xpath)
        try:
            element.clear()
        except WebDriverException:
            pass
        element.send_keys(value)

        return element
    
    def InputElement_2Values(xpath, value1, value2):
        '''• Usage: Send key with 2 values in input box
                return WebElement'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        try:
            element.clear()
        except WebDriverException:
            pass
        element.send_keys(value1)
        element.send_keys(value2)

        return element

    def Wait10s_InputElement(xpath, value):
        '''• Usage: Wait until the input box visible and send key value
                return WebElement'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.send_keys(value)

        return element
    
    def SwitchToFrame(frame_xpath):
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, frame_xpath)))
        frame = Commands.FindElement(frame_xpath)
        driver.switch_to.frame(frame)

        return frame
    
    def SwitchToDefaultContent():
        driver.switch_to.default_content()

    def ScrollDown():
        '''• Usuage: Scroll down, default height (0,-301)'''
        
        driver.execute_script("window.scrollTo(0,300)")
    
    def ScrollUp():
        '''• Usuage: Scroll down, default height (300,0)'''
        
        driver.execute_script("window.scrollTo(301, 0)")
    
    def Selectbox_ByValue(xpath, value):
        '''• Usage: Wait until select box is loaded
                select by value, return select box
                value = str()'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        Select(element).select_by_value(value)

        return element
    
    def Selectbox_ByIndex(xpath, index_number):
        '''• Usage: Wait until select box is loaded
                select by the index, return select box
                index_number = int()'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        Select(element).select_by_index(index_number)

        return element
    
    def Selectbox_ByVisibleText(xpath, selected_text):
        '''• Usage: Wait until select box is loaded
                select by visible text, return select box
                visible text = str()'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        Select(element).select_by_visible_text(selected_text)

        return element

    def MoveToElement(xpath):
        '''• Usage: Move to view element by ActionChains
                return WebElement'''

        element = driver.find_element_by_xpath(xpath)
        actions = ActionChains(driver)
        actions.move_to_element(element)
        actions.perform()
        time.sleep(1)

        return element

    def ReloadBrowser(page_xpath):
        driver.refresh()
        Waits.WaitUntilPageIsLoaded(page_xpath)

    def ActionsWithContainedXpath(action, xpath, replaced_value):
        '''Usage: action name can be used (wait10s / click / find'''

        element_xpath = Functions.xpath_ConvertXpath(xpath, replaced_value)
        if action == "wait10s":
            Waits.Wait10s_ElementLoaded(element_xpath)
        elif action == "click":
            Commands.ClickElement(element_xpath)
        elif action == "find":
            Commands.FindElement(element_xpath)

class Functions():
    def DefineCurrentURL():
        current_url = str(driver.current_url)

        return current_url

    def GetElementText(xpath):
        '''• Usage: Get and return element_text as str()'''

        element_text = str(driver.find_element_by_xpath(xpath).text)

        return element_text
    
    def GetInputValue(xpath):
        '''• Usage: Get and return input_value as str()
                 Use this function if element is input box'''

        input_element = driver.find_element_by_xpath(xpath)
        input_value = str(input_element.get_attribute("value"))

        return input_value
    
    def GetElementAttribute(xpath, attribute):
        '''• Usage: Get and return element_attribute as str()
                        (attribute can be value of 'class', 'style'... '''

        element = driver.find_element_by_xpath(xpath)
        element_attribute = str(element.get_attribute(attribute))

        return element_attribute

    def GetListLength(xpath):
        '''• Usage: Count how many elements are visible
                return a number int()'''

        list_length = int(len(driver.find_elements_by_xpath(xpath)))

        return list_length
    
    def xpath_ConvertXpath(xpath, replaced_value):
        '''• Usage: xpath which is being used must be written in style 'replaced_text'
                return str()'''

        if type(replaced_value) == int():
            '''It's used to define the order number of element
                        E.g: xpath + "[" + str(i) + "]" '''
                        # i=int()
            element_xpath = str(xpath).replace("order_number", str(replaced_value))
        
        elif type(replaced_value) == str():
            ''' It's used to replace the text in xpath
                        E.g: xpath = xpath + [contains(., 'replaced_text')] '''
                        # replaced_text=str()
            element_xpath = str(xpath).replace("replaced_text", str(replaced_value))

        return element_xpath

    def getRandomNumber_fromSpecificRange(first_number, last_number):
        '''• Usage: Get a list of random numbers
                return a number int()'''

        random_number = int(random.randint(first_number, last_number))

        return random_number

    def getRandomList_fromSpecificRange(picked_numbers, assigned_range):
        '''• Usage: Get a list of random numbers and remove duplicated number
                return a list()'''

        random_number = random(randint(range(assigned_range)))

        random_list = []
        i=1
        for i in range(assigned_range):
            random_number = random(randint(range(assigned_range)))
            random_list.append(random_number)
            
            random_list = list(dict.fromkeys(random_list))
            if len(random_list) == picked_numbers:
                break
            
            i+=1 

        return random_list

    def RemoveDuplicate_fromList(selected_list):
        '''• Usage: Remove duplicated items in the assigned list
                return the assigned list without duplicated item'''
        
        selected_list = list(dict.fromkeys(selected_list))

        return selected_list

    def checkIf_ElementVisible(xpath):
        '''• Usage: check element is visible
                    return True if element is visible'''
        
        try:
            driver.find_element_by_xpath(xpath)
            return True
        except WebDriverException:
            return False

    def waitIf_ElementVisible(xpath):
        '''• Usage: Wait 10s until element is visible
                    return True if element is visible'''
        
        try:
            Waits.Wait10s_ElementLoaded(xpath)
            return True
        except WebDriverException:
            return False

class Title():
    def FormatTitle(mail_title, send_date):
        title = "%s [Date: %s]" % (mail_title, send_date)

        return title
    
    def SplitTitle(mail_text):
        mail_title = str(mail_text).split(" [Date: ")[0]
        send_date = str(mail_text).split(" [Date: ")[1].replace("]", "")

        return (mail_title, send_date)

class SetUp():
    def Run():
        Logs.CreateLogFiles()
        Logs.ClearExcelFile()
        Files.ConfigFiles()

SetUp.Run()